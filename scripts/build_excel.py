#!/usr/bin/env python3
"""Build the final Excel report from the master dataset.

Contract (see .claude/skills/hf-excel/SKILL.md):
  stdout: {"rows":N,"excluded":N,"changed_rows":N,"output":"..."}
"""
import argparse
import json
import sys
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl.utils import get_column_letter

COLUMNS = [
    ("model_id", "모델 ID"),
    ("model_name", "모델명"),
    ("company_name", "회사명"),
    ("url", "URL"),
    ("technical_report_url", "기술 리포트 링크"),
    ("description", "모델설명"),
    ("domain", "특화 도메인"),
    ("modality", "지원 모달리티"),
    ("language", "지원 언어"),
    ("license", "라이선스"),
    ("release_date", "공개일"),
    ("last_modified_date", "최근 업데이트일"),
    ("parameters", "규모(파라미터)"),
    ("context_length", "컨텍스트 길이"),
    ("precision", "정밀도"),
    ("architecture", "모델 아키텍처"),
    ("construction_method", "모델 구축 방식"),
    ("base_model", "베이스 모델"),
    ("benchmark_language_knowledge", "벤치마크_언어지식이해"),
    ("benchmark_technical", "벤치마크_전문논리능력"),
    ("benchmark_multimodal", "벤치마크_멀티모달이해"),
    ("benchmark_safety", "벤치마크_안전성신뢰성"),
    ("benchmark_korean", "벤치마크_한국어"),
    ("benchmark_other", "벤치마크_기타"),
    ("knowledge_cutoff", "지식 최신성(일자)"),
]

DATE_COLUMNS = ["공개일", "최근 업데이트일"]
NUMERIC_COLUMNS = ["규모(파라미터)", "컨텍스트 길이"]


def load_master(path):
    records = []
    p = Path(path)
    if p.exists():
        for line in p.read_text(encoding="utf-8").splitlines():
            if line.strip():
                records.append(json.loads(line))
    return records


def to_row(rec):
    return {label: rec.get(key) for key, label in COLUMNS}


def autosize(ws, df):
    for i, col in enumerate(df.columns, start=1):
        if df.empty:
            max_len = len(str(col))
        else:
            max_len = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str).tolist()])
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 60)


def apply_column_types(df):
    """날짜 필드는 datetime으로, 숫자 필드는 numeric으로 캐스팅해 엑셀에 실제
    날짜/숫자 타입으로 기록되게 한다(문자열로 남으면 정렬·필터·수식에 안 걸림)."""
    for col in DATE_COLUMNS:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")
    for col in NUMERIC_COLUMNS:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


def format_numeric_columns(ws, df):
    """정수 필드(규모/컨텍스트 길이)가 12345.0처럼 소수점을 달고 보이지 않도록
    셀 표시 형식을 정수 콤마 포맷으로 지정한다."""
    for i, col in enumerate(df.columns, start=1):
        if col in NUMERIC_COLUMNS:
            letter = get_column_letter(i)
            for row in range(2, len(df) + 2):
                ws[f"{letter}{row}"].number_format = "#,##0"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--master-file", required=True)
    ap.add_argument("--changes-file", required=True)
    ap.add_argument("--output", required=True)
    args = ap.parse_args()

    records = load_master(args.master_file)
    if not records:
        print(json.dumps({"error": "master_dataset empty or missing"}, ensure_ascii=False))
        sys.exit(1)

    included = [r for r in records if not r.get("excluded")]
    excluded_count = len(records) - len(included)
    if not included:
        print(json.dumps({"error": "all records excluded, nothing to show"}, ensure_ascii=False))
        sys.exit(1)

    changes_path = Path(args.changes_file)
    changed_ids = {}
    if changes_path.exists():
        changes = json.loads(changes_path.read_text(encoding="utf-8"))
        changed_ids = {c["model_id"]: c["change_type"] for c in changes}

    all_df = pd.DataFrame([to_row(r) for r in included])
    all_df = apply_column_types(all_df)
    all_df.sort_values(by=["회사명", "최근 업데이트일"], ascending=[True, False], inplace=True, na_position="last")

    changes_rows = []
    for r in included:
        if r["model_id"] in changed_ids:
            row = {"변경유형": changed_ids[r["model_id"]]}
            row.update(to_row(r))
            changes_rows.append(row)
    changes_df = pd.DataFrame(changes_rows)
    if not changes_df.empty:
        changes_df = apply_column_types(changes_df)

    # 값이 없으면 그냥 빈 칸으로 둔다(null을 문자열로 채우면 날짜/숫자 컬럼의 실제
    # 타입이 깨져서 정렬·필터·수식에 안 걸리게 되므로, NaN/NaT를 그대로 to_excel에
    # 넘겨 openpyxl이 빈 셀로 쓰게 한다).
    output_path = Path(args.output)
    output_path = output_path.with_name(f"{output_path.stem}_{date.today().isoformat()}{output_path.suffix}")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl", date_format="YYYY-MM-DD", datetime_format="YYYY-MM-DD") as writer:
        all_df.to_excel(writer, sheet_name="All Models", index=False)
        changes_df.to_excel(writer, sheet_name="This Run Changes", index=False)
        autosize(writer.sheets["All Models"], all_df)
        autosize(writer.sheets["This Run Changes"], changes_df)
        format_numeric_columns(writer.sheets["All Models"], all_df)
        format_numeric_columns(writer.sheets["This Run Changes"], changes_df)

    print(json.dumps({
        "rows": len(included), "excluded": excluded_count,
        "changed_rows": len(changes_rows), "output": str(output_path),
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
